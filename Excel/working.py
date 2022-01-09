from openpyxl import workbook, load_workbook

wb = load_workbook('Excel/Grandes.xlsx')
ws = wb.active
ws['A2'].value = "test"
print(wb.sheetnames)

wb.save('Excel/Grandes.xlsx')